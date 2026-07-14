import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.client import BalanceType, Client
from app.models.user import User
from app.schemas.client import ClientCreate, ClientImportError, ClientImportResult, ClientUpdate


HEADER_ALIASES = {
    "client_name": {"company", "company name", "client name", "customer name", "name", "party name"},
    "address_l1": {"address l1", "address 1", "address line 1", "address"},
    "address_l2": {"address l2", "address 2", "address line 2"},
    "gstin": {"gstin", "gst no", "gst number", "gstin no", "gstin number"},
    "mobile": {"mobile", "phone", "phone no", "mobile no", "contact"},
    "email": {"email", "email id"},
    "opening_balance": {"opening balance", "opening_balance"},
    "balance_type": {"balance type", "balance_type", "dr cr", "debit credit"},
}


def _normalize_gstin(gstin: str | None) -> str | None:
    if gstin is None:
        return None
    normalized = gstin.strip().upper()
    return normalized or None


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").split()).strip()


def _normalize_header(value: Any) -> str:
    return _clean_cell(value).lower().replace("_", " ")


def _map_headers(headers: list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    normalized_headers = [_normalize_header(header) for header in headers]
    for field, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                mapped[field] = index
                break
    return mapped


def _normalize_row(row: tuple[Any, ...] | list[Any]) -> tuple[Any, ...]:
    values = tuple(row)
    non_empty = [_clean_cell(value) for value in values if _clean_cell(value)]
    if len(non_empty) == 1:
        value = non_empty[0]
        if "\t" in value:
            return tuple(value.split("\t"))
        if "," in value:
            return tuple(next(csv.reader([value])))
    return values


def _select_header_row(
    first_row: list[Any],
    rows: list[tuple[Any, ...]],
) -> tuple[dict[str, int], list[tuple[Any, ...]], int]:
    all_rows = [_normalize_row(first_row), *[_normalize_row(row) for row in rows]]
    required_headers = {"client_name", "gstin"}

    for index, row in enumerate(all_rows[:10]):
        headers = _map_headers(list(row))
        if required_headers.issubset(headers):
            return headers, all_rows[index + 1 :], index + 2

    best_headers = _map_headers(list(all_rows[0])) if all_rows else {}
    missing_headers = sorted(required_headers - set(best_headers))
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            f"Missing required columns: {', '.join(missing_headers)}. "
            "Expected columns like Company Name and GSTIN."
        ),
    )


def _cell(row: tuple[Any, ...], headers: dict[str, int], field: str) -> str:
    index = headers.get(field)
    if index is None or index >= len(row):
        return ""
    return _clean_cell(row[index])


def _parse_decimal(value: str, row_number: int, errors: list[ClientImportError]) -> Decimal:
    if not value:
        return Decimal("0.00")
    try:
        return Decimal(value.replace(",", ""))
    except InvalidOperation:
        errors.append(ClientImportError(row=row_number, message=f"Invalid opening balance: {value}"))
        return Decimal("0.00")


def _parse_balance_type(value: str) -> BalanceType:
    normalized = value.strip().lower()
    if normalized in {"credit", "cr", "c"}:
        return BalanceType.credit
    return BalanceType.debit


def get_client_by_id(db: Session, client_id: int) -> Client | None:
    return db.get(Client, client_id)


def get_client_by_gstin(db: Session, gstin: str | None) -> Client | None:
    normalized_gstin = _normalize_gstin(gstin)
    if normalized_gstin is None:
        return None
    return db.scalar(select(Client).where(Client.gstin == normalized_gstin))


def create_client(db: Session, payload: ClientCreate, current_user: User) -> Client:
    current_balance = payload.current_balance
    if current_balance is None:
        current_balance = payload.opening_balance

    client = Client(
        client_name=payload.client_name,
        billing_name=payload.billing_name,
        address=payload.address,
        gstin=_normalize_gstin(payload.gstin),
        mobile=payload.mobile,
        email=str(payload.email) if payload.email else None,
        opening_balance=payload.opening_balance,
        balance_type=payload.balance_type,
        current_balance=current_balance,
        created_by_user_id=current_user.id,
    )
    db.add(client)
    db.commit()
    db.refresh(client)
    return client


def _read_csv_rows(contents: bytes) -> tuple[list[Any], list[tuple[Any, ...]]]:
    try:
        text = contents.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = contents.decode("latin-1")

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(StringIO(text), dialect)
    try:
        headers = next(reader)
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV file is empty.") from None
    return headers, [tuple(row) for row in reader]


def _read_xlsx_rows(contents: bytes) -> tuple[list[Any], list[tuple[Any, ...]]]:
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read Excel file.",
        ) from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty.") from None
    return headers, list(rows)


async def import_clients_from_excel(db: Session, upload: UploadFile, current_user: User) -> ClientImportResult:
    filename = (upload.filename or "").lower()
    if not filename.endswith((".xlsx", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx Excel or .csv files are supported.",
        )

    contents = await upload.read()
    raw_headers, rows = _read_csv_rows(contents) if filename.endswith(".csv") else _read_xlsx_rows(contents)
    headers, data_rows, first_data_row_number = _select_header_row(raw_headers, rows)

    inserted = 0
    updated = 0
    skipped_duplicates = 0
    skipped_invalid = 0
    seen_gstins: set[str] = set()
    errors: list[ClientImportError] = []

    for row_number, row in enumerate(data_rows, start=first_data_row_number):
        client_name = _cell(row, headers, "client_name")
        gstin = _normalize_gstin(_cell(row, headers, "gstin"))
        if not client_name and not gstin:
            continue
        if not client_name:
            skipped_invalid += 1
            errors.append(ClientImportError(row=row_number, message="Client name is required."))
            continue
        if not gstin or len(gstin) != 15:
            skipped_invalid += 1
            errors.append(ClientImportError(row=row_number, message="Valid 15-character GSTIN is required."))
            continue
        if gstin in seen_gstins:
            skipped_duplicates += 1
            errors.append(ClientImportError(row=row_number, message=f"Duplicate GSTIN in file: {gstin}"))
            continue
        seen_gstins.add(gstin)

        address_parts = [_cell(row, headers, "address_l1"), _cell(row, headers, "address_l2")]
        address = ", ".join(part for part in address_parts if part) or None
        mobile = _cell(row, headers, "mobile") or None
        email = _cell(row, headers, "email") or None
        opening_balance = _parse_decimal(_cell(row, headers, "opening_balance"), row_number, errors)
        balance_type = _parse_balance_type(_cell(row, headers, "balance_type"))

        existing = get_client_by_gstin(db, gstin)
        if existing:
            existing.client_name = client_name
            existing.billing_name = client_name
            existing.address = address
            existing.mobile = mobile
            existing.email = email
            existing.opening_balance = opening_balance
            existing.balance_type = balance_type
            updated += 1
            continue

        db.add(
            Client(
                client_name=client_name,
                billing_name=client_name,
                address=address,
                gstin=gstin,
                mobile=mobile,
                email=email,
                opening_balance=opening_balance,
                balance_type=balance_type,
                current_balance=opening_balance,
                created_by_user_id=current_user.id,
            )
        )
        inserted += 1

    db.commit()
    return ClientImportResult(
        inserted=inserted,
        updated=updated,
        skipped_duplicates=skipped_duplicates,
        skipped_invalid=skipped_invalid,
        errors=errors[:50],
    )


def list_clients(db: Session, search: str | None = None) -> list[Client]:
    statement = select(Client).order_by(Client.client_name.asc())
    if search:
        statement = statement.where(Client.client_name.ilike(f"%{search.strip()}%"))
    return list(db.scalars(statement).all())


def update_client(db: Session, client: Client, payload: ClientUpdate) -> Client:
    update_data = payload.model_dump(exclude_unset=True)
    if "gstin" in update_data:
        update_data["gstin"] = _normalize_gstin(update_data["gstin"])
    if "email" in update_data and update_data["email"] is not None:
        update_data["email"] = str(update_data["email"])

    for field, value in update_data.items():
        setattr(client, field, value)

    db.commit()
    db.refresh(client)
    return client


def delete_client(db: Session, client: Client) -> None:
    db.delete(client)
    db.commit()
